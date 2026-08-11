"""regulatory_watch.changelog — ESMA's own published amendment history.

The ESMA message workbook ships two change-log sheets:

* ``Change log - Summary``  — one row per amendment: version, amendment date,
  trigger, explanation, amended documents, the RTS field codes affected, the
  reporting element, the XML path, the XSD amendment type and backward
  compatibility;
* ``Change log - XML path`` — per-amendment old XML path -> new XML path.

This is authority-published evidence of what really changed between Annex 2
source versions, and it is used here for exactly two things:

1. **Vocabulary coverage.** Every ESMA amendment type observed in the history
   is mapped onto the comparator's change-type vocabulary, so we can show the
   comparator is able to *express* the kinds of change ESMA has actually made.
2. **Historical baseline reconstruction.** A narrow, explicitly-bounded
   reverse-application of amendments that are fully specified by the change log
   (today: element additions, whose new paths are published in full). Anything
   not deterministically reversible is reported as a non-reconstructible
   amendment — never guessed, and never quietly skipped.

The change log is corroborating evidence. It never overrides a parsed value.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

from .annex2_spec import CODE_RE, SpecParseError
from .contracts import (
    ENUM_CHANGED,
    FIELD_ADDED,
    FIELD_REMOVED,
    MULTIPLICITY_CHANGED,
    NormalizedSpec,
    VALIDATION_RULE_CHANGED,
    XML_PATH_CHANGED,
)

SUMMARY_SHEET = "Change log - Summary"
XML_PATH_SHEET = "Change log - XML path"

#: ESMA's own "XSD amendment type" vocabulary -> our comparator change types.
#: Where ESMA states an amendment type we cannot express, the mapping is empty
#: and the amendment is reported as unmapped rather than silently dropped.
ESMA_AMENDMENT_TYPE_MAP: Dict[str, Tuple[str, ...]] = {
    "element tag addition": (FIELD_ADDED,),
    "element tag removal": (FIELD_REMOVED,),
    "element tag move": (XML_PATH_CHANGED,),
    "element tag move/amendment": (XML_PATH_CHANGED,),
    "element tag amendment": (XML_PATH_CHANGED,),
    "element path amendment": (XML_PATH_CHANGED,),
    "code change": (ENUM_CHANGED,),
    "element codes added": (ENUM_CHANGED,),
    "multiplicity change": (MULTIPLICITY_CHANGED,),
    "n/a": (VALIDATION_RULE_CHANGED,),   # XLS-only clarifications / rule edits
}

_CODE_SPLIT = re.compile(r"[\n|,;]+")


@dataclass
class ChangeLogEntry:
    """One ESMA-published amendment, verbatim plus a normalized view."""

    version: str
    amendment_id: str
    amendment_date: str
    amendment_trigger: str
    description: str
    amended_documents: str
    amended_xsd: str
    rts_field_codes: List[str]
    reporting_element: str
    xsd_amendment_types: List[str]
    backward_compatibility: str
    row_number: int
    xml_paths: List[Dict[str, str]] = field(default_factory=list)

    @property
    def annex2_codes(self) -> List[str]:
        return [c for c in self.rts_field_codes if CODE_RE.match(c)]

    @property
    def touches_annex2(self) -> bool:
        return bool(self.annex2_codes)

    @property
    def mapped_change_types(self) -> List[str]:
        out: List[str] = []
        for raw in self.xsd_amendment_types:
            for mapped in ESMA_AMENDMENT_TYPE_MAP.get(raw.strip().lower(), ()):
                if mapped not in out:
                    out.append(mapped)
        return out

    @property
    def unmapped_amendment_types(self) -> List[str]:
        return sorted({t for t in self.xsd_amendment_types
                       if t.strip().lower() not in ESMA_AMENDMENT_TYPE_MAP})

    def to_dict(self) -> Dict[str, Any]:
        return {
            "version": self.version,
            "amendment_id": self.amendment_id,
            "amendment_date": self.amendment_date,
            "amendment_trigger": self.amendment_trigger,
            "description": self.description,
            "amended_documents": self.amended_documents,
            "amended_xsd": self.amended_xsd,
            "rts_field_codes": list(self.rts_field_codes),
            "annex2_codes": self.annex2_codes,
            "reporting_element": self.reporting_element,
            "xsd_amendment_types": list(self.xsd_amendment_types),
            "mapped_change_types": self.mapped_change_types,
            "unmapped_amendment_types": self.unmapped_amendment_types,
            "backward_compatibility": self.backward_compatibility,
            "row_number": self.row_number,
            "xml_paths": list(self.xml_paths),
        }


def _cell(v: Any) -> str:
    if v is None:
        return ""
    text = str(v).strip()
    # openpyxl renders workbook dates as datetimes; keep the ISO date only.
    if len(text) == 19 and text[4] == "-" and text.endswith("00:00:00"):
        return text[:10]
    return text


def _split_multi(cell: str) -> List[str]:
    return [p.strip() for p in _CODE_SPLIT.split(cell) if p.strip()]


def extract_changelog(workbook_path: Path) -> List[ChangeLogEntry]:
    """Parse both change-log sheets into structured amendments.

    Raises :class:`SpecParseError` if the workbook cannot be opened or the
    summary sheet is absent.
    """
    import openpyxl

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise SpecParseError(f"workbook not found: {workbook_path}")
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True,
                                    data_only=True)
    except Exception as exc:              # noqa: BLE001
        raise SpecParseError(
            f"workbook could not be opened ({type(exc).__name__}): {exc}")

    try:
        if SUMMARY_SHEET not in wb.sheetnames:
            raise SpecParseError(
                f"workbook has no '{SUMMARY_SHEET}' sheet "
                f"(found: {list(wb.sheetnames)})")

        # -- XML path sheet, keyed by (version, amendment id) --------------- #
        paths: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
        if XML_PATH_SHEET in wb.sheetnames:
            ws = wb[XML_PATH_SHEET]
            for index, row in enumerate(ws.iter_rows(values_only=True)):
                if index == 0 or not row or not _cell(row[0]):
                    continue
                version = _cell(row[0])
                for amendment_id in _split_multi(_cell(row[1])) or [""]:
                    paths.setdefault((version, amendment_id), []).append({
                        "change_type": _cell(row[2]),
                        "description": _cell(row[3]),
                        "updated_xsd_message": _cell(row[4]),
                        "field_number": _cell(row[5]),
                        "old_xml_path": _cell(row[6]),
                        "new_xml_path": _cell(row[7]),
                        "row_number": index + 1,
                    })

        # -- summary sheet --------------------------------------------------- #
        entries: List[ChangeLogEntry] = []
        ws = wb[SUMMARY_SHEET]
        last_version = ""
        last_date = ""
        last_trigger = ""
        for index, row in enumerate(ws.iter_rows(values_only=True)):
            if index == 0 or not row:
                continue
            version = _cell(row[0]) or last_version
            date = _cell(row[1]) or last_date
            trigger = _cell(row[2]) or last_trigger
            last_version, last_date, last_trigger = version, date, trigger
            amendment_id = _cell(row[4])
            description = _cell(row[5])
            if not version or not (amendment_id or description):
                continue
            entries.append(ChangeLogEntry(
                version=version,
                amendment_id=amendment_id,
                amendment_date=date,
                amendment_trigger=trigger,
                description=description,
                amended_documents=_cell(row[6]),
                amended_xsd=_cell(row[7]),
                rts_field_codes=_split_multi(_cell(row[8])),
                reporting_element=_cell(row[9]),
                xsd_amendment_types=_split_multi(_cell(row[11])),
                backward_compatibility=_cell(row[12]),
                row_number=index + 1,
                xml_paths=paths.get((version, amendment_id), []),
            ))
    finally:
        try:
            wb.close()
        except Exception:                 # noqa: BLE001
            pass

    return entries


# --------------------------------------------------------------------------- #
# Historical baseline reconstruction
# --------------------------------------------------------------------------- #

@dataclass
class Reconstruction:
    """The result of reverse-applying an ESMA amendment to a parsed spec."""

    spec: NormalizedSpec
    applied: List[Dict[str, Any]] = field(default_factory=list)
    not_reconstructible: List[Dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "spec_version": self.spec.spec_version,
            "applied": list(self.applied),
            "not_reconstructible": list(self.not_reconstructible),
        }


def _paths_added_by(entry: ChangeLogEntry) -> List[str]:
    out: List[str] = []
    for row in entry.xml_paths:
        if "addition" not in row.get("change_type", "").lower():
            continue
        new_path = row.get("new_xml_path", "")
        if not new_path or new_path.lower() == "removed":
            continue
        # "(currency)" annotates an attribute of the preceding element path.
        new_path = new_path.split(" (")[0].strip()
        if new_path and new_path not in out:
            out.append(new_path)
    return out


def reconstruct_prior_spec(spec: NormalizedSpec,
                           entries: Sequence[ChangeLogEntry],
                           target_version: str,
                           prior_version_label: str) -> Reconstruction:
    """Reverse-apply the amendments introduced in ``target_version``.

    Deliberately narrow. Only an *element addition* whose new XML paths the
    change log publishes in full is reversible from the published evidence: the
    field it introduced is removed from the spec. Every other amendment type is
    recorded in ``not_reconstructible`` with the reason, because the change log
    does not publish the pre-amendment value and inventing one would fabricate
    a regulatory difference.

    The comparator is not involved and is not modified by this function.
    """
    prior = NormalizedSpec.from_dict(spec.to_dict())
    prior.spec_version = prior_version_label
    applied: List[Dict[str, Any]] = []
    blocked: List[Dict[str, Any]] = []

    for entry in entries:
        if entry.version != target_version:
            continue
        added_paths = _paths_added_by(entry)
        # Reversible only when the amendment is PURELY an element addition and
        # ESMA published the new paths in full. A mixed amendment (an addition
        # bundled with a move or a code change) is not reversible from the
        # change log alone.
        types = [t.strip().lower() for t in entry.xsd_amendment_types]
        reversible = bool(added_paths) and bool(types) and all(
            "addition" in t for t in types)
        if not reversible:
            blocked.append({
                "version": entry.version,
                "amendment_id": entry.amendment_id,
                "description": entry.description,
                "xsd_amendment_types": entry.xsd_amendment_types,
                "annex2_codes": entry.annex2_codes,
                "reason": ("the change log does not publish the "
                           "pre-amendment value for this amendment type; a "
                           "prior-version baseline cannot be reconstructed "
                           "without the earlier source artefact"),
            })
            continue

        removed: List[str] = []
        for code in sorted(prior.fields):
            field_obj = prior.fields[code]
            for path in added_paths:
                if (field_obj.xml_path
                        and (path == field_obj.xml_path
                             or path.startswith(field_obj.xml_path + "/"))):
                    removed.append(code)
                    break
        for code in removed:
            prior.fields.pop(code, None)
            if code in prior.order:
                prior.order.remove(code)

        applied.append({
            "version": entry.version,
            "amendment_id": entry.amendment_id,
            "description": entry.description,
            "xsd_amendment_types": entry.xsd_amendment_types,
            "added_paths": added_paths,
            "codes_removed_from_prior_baseline": sorted(removed),
            "in_annex2_scope": bool(removed),
        })

    return Reconstruction(spec=prior, applied=applied,
                          not_reconstructible=blocked)


def vocabulary_coverage(entries: Sequence[ChangeLogEntry]
                        ) -> Dict[str, Any]:
    """How much of ESMA's published amendment vocabulary we can express."""
    seen: Dict[str, int] = {}
    unmapped: Dict[str, int] = {}
    annex2_entries = 0
    for entry in entries:
        if entry.touches_annex2:
            annex2_entries += 1
        for raw in entry.xsd_amendment_types:
            key = raw.strip().lower()
            seen[key] = seen.get(key, 0) + 1
            if key not in ESMA_AMENDMENT_TYPE_MAP:
                unmapped[key] = unmapped.get(key, 0) + 1
    return {
        "entries": len(entries),
        "entries_touching_annex2_codes": annex2_entries,
        "esma_amendment_types": dict(sorted(seen.items())),
        "unmapped_amendment_types": dict(sorted(unmapped.items())),
    }
