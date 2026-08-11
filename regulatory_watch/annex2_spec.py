"""regulatory_watch.annex2_spec — deterministic Annex 2 spec normalization.

Turns the authoritative ESMA artefacts into a :class:`NormalizedSpec`: the
minimum normalized representation needed to compare two Annex 2 versions. This
is **not** a generic regulation framework — it models exactly the Annex 2
attributes that can change what Trakt must report.

Two authoritative artefacts, each used for what only it can supply:

* the **ESMA message/mapping workbook** (``DRAFT1auth.099.001.04 ... .xlsx``)
  gives, per RTS field code: field name, message-item definition, XML tag, the
  full XML ``PATH``, multiplicity, the data-type / code-list *name*, the
  pattern, and the validation-rule text;
* the **XSD** resolves those code-list names into actual values — the
  enumeration for a coded field, and the
  ``NoDataAllowedJustification{1,2,3,4}Code`` vocabularies that state exactly
  which ND options a field permits.

Derivation rules, all deterministic:

======================  =====================================================
attribute               derivation
======================  =====================================================
``xml_path``            the element node: the unique shortest workbook PATH for
                        the code that prefixes every other PATH for that code
``value_paths``         PATHs under the element node, excluding NoDataOptn
``data_type``/          the first value leaf (in sheet order) carrying a
``format_pattern``      DATA TYPE / Pattern
``nd_allowed``          XSD enumeration of the justification type on
                        ``<element>/NoDataOptn/NoData``; ``[]`` when the
                        workbook publishes no NoDataOptn branch
``enum_values``         XSD enumeration of the value leaf's code-list type
``mandatory``           minimum occurrence of the element node's multiplicity
``validation_rules``    verbatim ``Validation rule`` cells, de-duplicated in
                        sheet order
``order_index``         position of the code in workbook sheet order
======================  =====================================================

Nothing is inferred. Where an attribute cannot be derived it is set to
``UNKNOWN`` and named in :attr:`SpecField.unresolved`, and the comparator
downgrades any delta touching it to ``review-required``. No LLM is involved at
any point.

Scope: the Annex 2 residential-real-estate **performing** branch
(``ResdtlRealEsttLn/PrfrmgLn``), i.e. the same branch the live Trakt Annex 2
pathway targets. Cancellation (``/Cxl/``) rows are excluded.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from . import PARSER_VERSION, REGIME
from .contracts import (
    NormalizedSpec,
    Provenance,
    SourceSnapshot,
    SpecField,
    UNKNOWN,
)

# --------------------------------------------------------------------------- #
# Workbook layout (verified against DRAFT1auth.099.001.04 v1.3.1)
# --------------------------------------------------------------------------- #

DEFAULT_SHEET = "DRAFT1auth.099.001.04"

#: Header row of the message-item table, 0-based. Row 0-2 are the title block.
HEADER_ROW_INDEX = 3

#: Columns the normalizer requires. A workbook missing any of these has an
#: unexpected structure and is refused rather than partially parsed.
REQUIRED_COLUMNS = (
    "RTS Field code", "RTS Field name", "XML TAG", "PATH", "MULTIPLICITY",
    "Template", "Performing/Non Performing", "DATA TYPE / CODE", "Pattern",
    "Validation rule", "MESSAGE ITEM DEFINITION",
)

OPTIONAL_COLUMNS = ("Field code", "Item type", "CODE VALUE")

#: Annex 2 codes: RREL* (underlying exposure) and RREC* (collateral).
CODE_RE = re.compile(r"^(RREL|RREC)\d+$")
_CODE_SPLIT_RE = re.compile(r"[\n|,;]+")

#: Template / performance scope of the live Trakt Annex 2 pathway.
DEFAULT_TEMPLATES = ("ALL", "RRE")
DEFAULT_PERFORMANCE = "PRF"
_PERF_ANY = "PRF/NPRF"

_ND_OPT = "NoDataOptn"
_MULT_RE = re.compile(r"^\[?(\d+)\.\.(\d+|n|N)\]?$")


class SpecParseError(ValueError):
    """The source artefact could not be parsed into an Annex 2 spec.

    Raised for a corrupt workbook, a missing sheet or an unexpected column
    layout. Callers surface this as an undetermined comparison, never as
    "no regulatory change".
    """


# --------------------------------------------------------------------------- #
# Raw extraction (the only I/O in this module)
# --------------------------------------------------------------------------- #

@dataclass(frozen=True)
class RawRow:
    """One workbook message-item row, verbatim, with its sheet coordinates."""

    row_number: int                  # 1-based sheet row, for provenance
    codes: Tuple[str, ...]           # RREL/RREC codes named by this row
    field_code: str                  # ESMA numeric "Field code"
    item_type: str                   # "E" element / "P" parent
    template: str
    performance: str
    rts_field_name: str
    xml_tag: str
    path: str
    multiplicity: str
    data_type: str
    pattern: str
    validation_rule: str
    definition: str

    def to_dict(self) -> Dict[str, Any]:
        d = {k: getattr(self, k) for k in self.__dataclass_fields__}
        d["codes"] = list(self.codes)
        return d

    @staticmethod
    def from_dict(data: Dict[str, Any]) -> "RawRow":
        data = dict(data)
        data["codes"] = tuple(data.get("codes") or ())
        known = set(RawRow.__dataclass_fields__)
        return RawRow(**{k: v for k, v in data.items() if k in known})


def _cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _split_codes(cell: str) -> Tuple[str, ...]:
    out = []
    for part in _CODE_SPLIT_RE.split(cell):
        part = part.strip()
        if CODE_RE.match(part) and part not in out:
            out.append(part)
    return tuple(out)


def _clean_tag(cell: str) -> str:
    return cell.strip().strip("<>").strip()


def extract_rows(workbook_path: Path, sheet_name: str = DEFAULT_SHEET,
                 templates: Sequence[str] = DEFAULT_TEMPLATES,
                 performance: str = DEFAULT_PERFORMANCE,
                 include_cancellation: bool = False) -> List[RawRow]:
    """Stream the workbook and return the in-scope Annex 2 rows.

    Raises :class:`SpecParseError` for a corrupt file, a missing sheet or a
    missing required column — an unexpected source structure must be loud.
    """
    import openpyxl                       # local: keeps import cost off callers

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise SpecParseError(f"workbook not found: {workbook_path}")
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True,
                                    data_only=True)
    except Exception as exc:              # noqa: BLE001 - any corruption
        raise SpecParseError(
            f"workbook could not be opened ({type(exc).__name__}): {exc}")

    try:
        if sheet_name not in wb.sheetnames:
            raise SpecParseError(
                f"workbook has no sheet '{sheet_name}' "
                f"(found: {list(wb.sheetnames)})")
        ws = wb[sheet_name]

        header: Dict[str, int] = {}
        rows: List[RawRow] = []
        wanted_templates = {t.upper() for t in templates}
        perf = performance.upper().strip()

        for index, values in enumerate(ws.iter_rows(values_only=True)):
            if index < HEADER_ROW_INDEX:
                continue
            if index == HEADER_ROW_INDEX:
                header = {_cell(v): j for j, v in enumerate(values) if _cell(v)}
                missing = [c for c in REQUIRED_COLUMNS if c not in header]
                if missing:
                    raise SpecParseError(
                        f"workbook sheet '{sheet_name}' is missing required "
                        f"columns: {missing}")
                continue

            def col(name: str) -> str:
                j = header.get(name)
                if j is None or j >= len(values):
                    return ""
                return _cell(values[j])

            path = col("PATH")
            if not path:
                continue
            # Cancellation branch: excluded as a segment, so both '/Cxl/...'
            # and a bare '.../Cxl' leaf are dropped.
            if not include_cancellation and "Cxl" in _split_path(path):
                continue
            template = col("Template").upper()
            if template not in wanted_templates:
                continue
            pnp = col("Performing/Non Performing").upper()
            if pnp and pnp not in {_PERF_ANY, perf}:
                continue
            # Untagged container rows are KEPT: a code's element node is often
            # an untagged parent, and its multiplicity is only published there.
            codes = _split_codes(col("RTS Field code"))

            rows.append(RawRow(
                row_number=index + 1,          # 1-based, as shown in Excel
                codes=codes,
                field_code=col("Field code"),
                item_type=col("Item type"),
                template=template,
                performance=pnp,
                rts_field_name=col("RTS Field name"),
                xml_tag=_clean_tag(col("XML TAG")),
                path=path,
                multiplicity=col("MULTIPLICITY"),
                data_type=col("DATA TYPE / CODE"),
                pattern=col("Pattern"),
                validation_rule=col("Validation rule"),
                definition=col("MESSAGE ITEM DEFINITION"),
            ))
    finally:
        try:
            wb.close()
        except Exception:                  # noqa: BLE001 - best effort close
            pass

    if not rows:
        raise SpecParseError(
            f"no Annex 2 rows found in '{sheet_name}' for templates "
            f"{sorted(wanted_templates)} / performance '{perf}'")
    return rows


def extract_codelists(xsd_path: Path) -> Dict[str, List[str]]:
    """Return ``{simpleType name: [enumeration values]}`` from the ESMA schema.

    Values are returned in schema order; the normalizer sorts them, because the
    order of an enumeration is not itself a reporting requirement.
    """
    from lxml import etree

    xsd_path = Path(xsd_path)
    if not xsd_path.exists():
        raise SpecParseError(f"schema not found: {xsd_path}")
    try:
        root = etree.parse(str(xsd_path)).getroot()
    except Exception as exc:              # noqa: BLE001 - any corruption
        raise SpecParseError(
            f"schema could not be parsed ({type(exc).__name__}): {exc}")

    ns = "{http://www.w3.org/2001/XMLSchema}"
    out: Dict[str, List[str]] = {}
    for simple_type in root.iter(f"{ns}simpleType"):
        name = simple_type.get("name")
        if not name:
            continue
        values = [e.get("value") for e in simple_type.iter(f"{ns}enumeration")
                  if e.get("value") is not None]
        if values:
            out[name] = values
    return out


def schema_namespace(xsd_path: Path) -> str:
    """Target namespace of the schema, or ``UNKNOWN``."""
    from lxml import etree
    try:
        return etree.parse(str(xsd_path)).getroot().get(
            "targetNamespace") or UNKNOWN
    except Exception:                     # noqa: BLE001
        return UNKNOWN


# --------------------------------------------------------------------------- #
# Normalization (pure — no I/O, so fixtures replay exactly)
# --------------------------------------------------------------------------- #

def _split_path(path: str) -> List[str]:
    return [p for p in path.strip("/").split("/") if p]


def _min_occurs(multiplicity: str) -> Optional[int]:
    m = _MULT_RE.match(multiplicity.replace(" ", ""))
    if not m:
        return None
    try:
        return int(m.group(1))
    except ValueError:
        return None


def _is_under(path: str, parent: str) -> bool:
    return path == parent or path.startswith(parent.rstrip("/") + "/")


def _common_ancestor(paths: Sequence[str]) -> str:
    """Longest common ancestor path of ``paths``, as a '/'-prefixed string.

    Used to locate a code's element node. The workbook does not always tag the
    element node row itself with the RTS code (some codes are tagged only on
    their value/NoData leaves), so the element node is derived from the leaves
    rather than assumed to be one of them.
    """
    if not paths:
        return ""
    parts = [_split_path(p) for p in paths]
    common: List[str] = []
    for segments in zip(*parts):
        first = segments[0]
        if all(s == first for s in segments):
            common.append(first)
        else:
            break
    return "/" + "/".join(common) if common else ""


def _dedupe(values: Iterable[str]) -> List[str]:
    seen, out = set(), []
    for v in values:
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


def normalize(rows: Sequence[RawRow], codelists: Dict[str, List[str]],
              *, artefact_id: str, schema_artefact_id: str,
              sheet_name: str = DEFAULT_SHEET, spec_version: str = UNKNOWN,
              scope: Optional[Dict[str, Any]] = None,
              snapshots: Optional[Sequence[SourceSnapshot]] = None,
              parser_version: str = PARSER_VERSION) -> NormalizedSpec:
    """Build a :class:`NormalizedSpec` from extracted rows + XSD code lists.

    Pure and total: any code the rows cannot resolve is emitted with the
    offending attributes ``UNKNOWN`` and listed in ``unresolved``, plus a
    structured parse warning. Nothing is dropped silently.
    """
    grouped: Dict[str, List[RawRow]] = {}
    for row in rows:
        for code in row.codes:
            grouped.setdefault(code, []).append(row)

    warnings: List[Dict[str, Any]] = []
    fields: Dict[str, SpecField] = {}

    # The report anchor is the common ancestor of every in-scope row — the
    # container the whole message hangs from. A code whose own element node is
    # not a STRICT descendant of it spans disjoint parts of the message tree,
    # which the normalizer refuses to resolve.
    #
    # NOTE for fixture authors: this is relative to the rows passed in. A slice
    # narrow enough that one code's element node IS the slice's own anchor will
    # read as ambiguous. Keep a shallow code (e.g. RREL1) in any slice.
    report_anchor = _common_ancestor([r.path for r in rows])
    anchor_depth = len(_split_path(report_anchor))

    # Path -> row for every in-scope row. A code's element node is often an
    # untagged container row, so its multiplicity is read from here rather than
    # from the code's own (leaf) rows.
    by_path: Dict[str, RawRow] = {}
    for row in sorted(rows, key=lambda r: r.row_number):
        by_path.setdefault(row.path, row)

    # Source sequence: codes ordered by their first appearance in the sheet.
    ordered_codes = sorted(grouped,
                           key=lambda c: min(r.row_number
                                             for r in grouped[c]))

    for order_index, code in enumerate(ordered_codes):
        code_rows = sorted(grouped[code],
                           key=lambda r: (len(_split_path(r.path)),
                                          r.row_number))
        unresolved: List[str] = []
        provenance = [
            Provenance(artefact_id=artefact_id,
                       locator=f"sheet={sheet_name};row={r.row_number}",
                       detail=r.xml_tag or r.path)
            for r in code_rows
        ]

        # -- value branch vs ND branch -------------------------------------- #
        # The reportable element is where a VALUE goes, so the element node is
        # derived from the value rows only. The workbook publishes the ND
        # branch under a sibling path for a handful of codes (RREC8, RREL80);
        # deriving the element node from the value rows is stable under that.
        nd_rows = [r for r in code_rows if _ND_OPT in _split_path(r.path)]
        nd_row_numbers = {r.row_number for r in nd_rows}
        value_rows = [r for r in code_rows
                      if r.row_number not in nd_row_numbers]

        element_path = _common_ancestor(
            [r.path for r in (value_rows or code_rows)])
        ambiguous = len(_split_path(element_path)) <= anchor_depth
        if ambiguous:
            # The code's value rows span disjoint parts of the message tree:
            # the workbook is ambiguous, or the code has been reused. Never
            # pick one branch silently.
            unresolved.extend(["xml_path", "xml_tag", "multiplicity",
                               "mandatory"])
            warnings.append({
                "code": code,
                "warning": "AMBIGUOUS_ELEMENT_NODE",
                "detail": ("more than one disjoint XML path is published for "
                           "this code in scope"),
                "paths": sorted({r.path for r in code_rows}),
            })

        element_row = by_path.get(element_path)

        # -- ND permission -------------------------------------------------- #
        # A field may publish more than one NoData leaf: the primary
        # NoDataOptn/NoData plus the nested NoDataOptn/NoData4/NoData branch
        # that carries ND4. The permitted set is the union of every published
        # justification vocabulary — that is what the schema actually allows.
        nd_allowed: Optional[List[str]]
        nd_types: List[str] = []
        nd_path = UNKNOWN
        nd_roots = sorted({
            "/".join(_split_path(r.path)[:_split_path(r.path).index(_ND_OPT) + 1])
            for r in nd_rows
        }, key=lambda p: (len(_split_path(p)), p))
        if nd_roots:
            nd_path = "/" + nd_roots[0]
        nd_leaves = [r for r in nd_rows
                     if _split_path(r.path)[-1] == "NoData" and r.data_type]
        if nd_rows and not nd_leaves:
            # A NoDataOptn branch with no typed NoData leaf: a structure the
            # normalizer does not model. Do NOT conclude "ND not permitted".
            nd_allowed = None
            unresolved.append("nd_allowed")
            warnings.append({
                "code": code,
                "warning": "UNSUPPORTED_ND_STRUCTURE",
                "detail": "NoDataOptn branch present without a typed NoData leaf",
            })
        elif nd_leaves:
            permitted: set = set()
            unknown_types: List[str] = []
            for leaf in sorted(nd_leaves, key=lambda r: r.row_number):
                if leaf.data_type not in nd_types:
                    nd_types.append(leaf.data_type)
                values = codelists.get(leaf.data_type)
                if values is None:
                    unknown_types.append(leaf.data_type)
                    continue
                permitted.update(values)
                provenance.append(Provenance(
                    artefact_id=schema_artefact_id,
                    locator=f"xsd:simpleType/{leaf.data_type}",
                    detail="ND justification vocabulary"))
            if unknown_types:
                nd_allowed = None
                unresolved.append("nd_allowed")
                warnings.append({
                    "code": code,
                    "warning": "UNKNOWN_ND_JUSTIFICATION_TYPE",
                    "detail": ("the schema publishes no enumeration for "
                               f"{sorted(set(unknown_types))}"),
                })
            else:
                nd_allowed = sorted(permitted)
        else:
            # No NoDataOptn branch published: ND is deterministically not
            # permitted for this field.
            nd_allowed = []

        if nd_roots and not ambiguous and not _is_under("/" + nd_roots[0],
                                                        element_path):
            # Real inconsistency in the ESMA workbook: the ND branch is a
            # sibling of the value branch rather than a child of the element.
            # Recorded as evidence; it does not block ND derivation.
            warnings.append({
                "code": code,
                "warning": "ND_BRANCH_PATH_INCONSISTENT",
                "detail": (f"ND branch '/{nd_roots[0]}' is not nested under "
                           f"the value element '{element_path}'"),
            })

        # -- value leaves ---------------------------------------------------- #
        typed = [r for r in sorted(value_rows, key=lambda r: r.row_number)
                 if r.data_type]
        primary = typed[0] if typed else None

        data_type = primary.data_type if primary else UNKNOWN
        format_pattern = (primary.pattern or UNKNOWN) if primary else UNKNOWN
        if primary is None:
            unresolved.extend(["data_type", "format_pattern"])
            warnings.append({
                "code": code,
                "warning": "NO_TYPED_VALUE_LEAF",
                "detail": "no value leaf publishes a DATA TYPE / CODE",
            })

        # -- enumeration ------------------------------------------------------ #
        enum_type = UNKNOWN
        enum_values: Optional[List[str]] = None
        if primary is not None and primary.data_type in codelists:
            enum_type = primary.data_type
            enum_values = sorted(codelists[primary.data_type])
            provenance.append(Provenance(
                artefact_id=schema_artefact_id,
                locator=f"xsd:simpleType/{enum_type}",
                detail="allowed values"))
        elif primary is not None and primary.pattern.strip().lower() == "enumeration":
            # The workbook says the leaf is coded but the schema publishes no
            # enumeration under that name — report it, do not invent values.
            enum_type = primary.data_type or UNKNOWN
            unresolved.append("enum_values")
            warnings.append({
                "code": code,
                "warning": "UNRESOLVED_ENUMERATION",
                "detail": (f"workbook declares an enumeration but the schema "
                           f"has no simpleType '{primary.data_type}'"),
            })

        # -- multiplicity / mandatory ---------------------------------------- #
        raw_multiplicity = element_row.multiplicity if element_row else ""
        multiplicity = raw_multiplicity or UNKNOWN
        min_occ = _min_occurs(raw_multiplicity)
        if ambiguous:
            mandatory = UNKNOWN
        elif min_occ is None:
            mandatory = UNKNOWN
            for attr in ("mandatory", "multiplicity"):
                if attr not in unresolved:
                    unresolved.append(attr)
            warnings.append({
                "code": code,
                "warning": "UNPARSEABLE_MULTIPLICITY",
                "detail": (f"multiplicity cell '{raw_multiplicity}'"
                           if element_row else
                           "no workbook row publishes the element node"),
            })
        else:
            mandatory = "mandatory" if min_occ >= 1 else "optional"

        # -- labels / description --------------------------------------------- #
        label = next((r.rts_field_name for r in code_rows
                      if r.rts_field_name), UNKNOWN)
        description = next((r.definition for r in code_rows
                            if r.definition), UNKNOWN)
        if label == UNKNOWN:
            unresolved.append("label")
        if description == UNKNOWN:
            unresolved.append("description")

        fields[code] = SpecField(
            code=code,
            record_group=code[:4],
            label=label,
            description=description,
            data_type=data_type,
            format_pattern=format_pattern,
            mandatory=mandatory,
            multiplicity=multiplicity,
            nd_allowed=nd_allowed,
            nd_justification_type=("+".join(nd_types) if nd_types else UNKNOWN),
            enum_type=enum_type,
            enum_values=enum_values,
            xml_tag=(UNKNOWN if ambiguous
                     else (_split_path(element_path)[-1] or UNKNOWN)),
            xml_path=(UNKNOWN if ambiguous else element_path),
            value_paths=sorted({r.path for r in value_rows}),
            nd_path=nd_path,
            validation_rules=_dedupe(
                r.validation_rule for r in
                sorted(code_rows, key=lambda r: r.row_number)),
            order_index=order_index,
            provenance=provenance,
            unresolved=sorted(set(unresolved)),
        )

    return NormalizedSpec(
        regime=REGIME,
        spec_version=spec_version,
        parser_version=parser_version,
        fields=fields,
        order=list(ordered_codes),
        snapshots=list(snapshots or []),
        scope=dict(scope or {}),
        parse_warnings=sorted(
            warnings, key=lambda w: (str(w.get("code")), str(w.get("warning")))),
    )


# --------------------------------------------------------------------------- #
# Convenience: artefacts -> spec
# --------------------------------------------------------------------------- #

def parse_annex2_spec(workbook_path: Path, xsd_path: Path, *,
                      workbook_artefact_id: str = "esma_annex2_message_workbook",
                      schema_artefact_id: str = "esma_annex2_xml_schema",
                      sheet_name: str = DEFAULT_SHEET,
                      templates: Sequence[str] = DEFAULT_TEMPLATES,
                      performance: str = DEFAULT_PERFORMANCE,
                      spec_version: str = UNKNOWN,
                      snapshots: Optional[Sequence[SourceSnapshot]] = None,
                      ) -> NormalizedSpec:
    """Parse the two authoritative artefacts into one normalized spec."""
    rows = extract_rows(workbook_path, sheet_name=sheet_name,
                        templates=templates, performance=performance)
    codelists = extract_codelists(xsd_path)
    scope = {
        "sheet": sheet_name,
        "templates": list(templates),
        "performance": performance,
        "cancellation_rows": "excluded",
        "asset_branch": "ResdtlRealEsttLn/PrfrmgLn",
        "schema_namespace": schema_namespace(xsd_path),
    }
    return normalize(rows, codelists, artefact_id=workbook_artefact_id,
                     schema_artefact_id=schema_artefact_id,
                     sheet_name=sheet_name, spec_version=spec_version,
                     scope=scope, snapshots=snapshots)
