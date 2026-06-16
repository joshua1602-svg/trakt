#!/usr/bin/env python3
"""build_annex2_universe.py
==========================

Derive the authoritative ESMA Annex 2 field universe config from the workbook
template and write it as a diffable YAML the onboarding workflow reads at
runtime (no openpyxl dependency in the hot path).

The workbook is the ESMA Annex 2 template (one row per field): field code,
field name, content to report, ND1-ND4 / ND5 eligibility and format. It is the
authoritative target universe for Annex 2 coverage.

Usage::

    python scripts/build_annex2_universe.py \
        [config/regime/annex2_template_workbook.xlsx] \
        [config/regime/annex2_field_universe.yaml]
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import yaml

_REPO_ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_XLSX = _REPO_ROOT / "config" / "regime" / "annex2_template_workbook.xlsx"
_DEFAULT_YAML = _REPO_ROOT / "config" / "regime" / "annex2_field_universe.yaml"

_CODE_RE = re.compile(r"^(RREL|RREC)\d+$")
# Column indices in the template (0-based), per the ESMA Annex 2 template layout.
_C_SECTION = 1
_C_CODE = 2
_C_NAME = 3
_C_CONTENT = 4
_C_ND14 = 7
_C_ND5 = 8
_C_FORMAT = 9


def _yes(v) -> bool:
    return str(v or "").strip().upper() in ("YES", "Y", "TRUE")


def build(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    fields: dict = {}
    for row in ws.iter_rows(values_only=True):
        code = str(row[_C_CODE]).strip() if row[_C_CODE] is not None else ""
        if not _CODE_RE.match(code) or code in fields:
            continue
        fields[code] = {
            "field_name": str(row[_C_NAME] or "").strip(),
            "section": str(row[_C_SECTION] or "").strip(),
            "content": str(row[_C_CONTENT] or "").strip(),
            "nd1_4_allowed": _yes(row[_C_ND14]),
            "nd5_allowed": _yes(row[_C_ND5]),
            "format": str(row[_C_FORMAT] or "").strip(),
        }
    return {
        "regime": "ESMA_Annex2",
        "source": xlsx_path.name,
        "generated_by": "scripts/build_annex2_universe.py",
        "note": ("Authoritative workbook-derived ESMA Annex 2 field universe. "
                 "One entry per Annex 2 field code. Do not edit by hand — "
                 "regenerate from the template workbook."),
        "field_count": len(fields),
        "fields": fields,
    }


def main(argv=None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    xlsx = Path(argv[0]) if len(argv) > 0 else _DEFAULT_XLSX
    out = Path(argv[1]) if len(argv) > 1 else _DEFAULT_YAML
    data = build(xlsx)
    out.write_text(yaml.safe_dump(data, sort_keys=False, allow_unicode=True),
                   encoding="utf-8")
    print(f"Wrote {data['field_count']} Annex 2 fields -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
